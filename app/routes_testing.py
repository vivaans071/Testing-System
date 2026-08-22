"""
Testing sample inventory: Department -> Campaign -> Lot -> Mini.

Kept in its own router (prefix /api/testing) rather than bolted onto
routes_admin, because it's a separate tab with its own data model and mixing
it in would push that file past 2,000 lines.

The shape mirrors how the source spreadsheet actually worked:
  * results were recorded per LOT, not per mini (minis only had notes)
  * the test columns differ by sample type (Minis track FTT/HDT/StepC,
    Floaters track Hot Poling/STT/FTT), so they can't be hardcoded columns
  * cells weren't booleans - they held things like "HDT Skipped" and
    "Not signed off but completed" - so each test carries a status
"""
import io
import json
import re
from datetime import date as _date, datetime as _dt
from datetime import date as date_type

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from . import models
from .database import get_db
from .security import current_user, require_admin

router = APIRouter(prefix="/api/testing", tags=["testing"])

# A test cell is simply checked or not. The earlier five-state cycle turned out
# to be more bookkeeping than anyone wanted. "" (unchecked) isn't stored.
STATES = ["", "done"]


def _as_text(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None                     # a checkbox isn't a description
    return str(v).strip() or None


def _truthy(v):
    """Excel checkmarks arrive as booleans, but people also type x / yes /
    a date. Anything present and not explicitly negative counts as done."""
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (_dt, _date)):
        return True
    t = str(v).strip().lower()
    return bool(t) and t not in ("0", "no", "n", "false", "-", "\u2013", "na", "n/a")


def _parse_date_loose(v):
    if v is None:
        return None
    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, _date):
        return v
    try:
        return _date.fromisoformat(str(v).strip()[:10])
    except ValueError:
        return None


def _parse_date(value):
    if not value:
        return None
    try:
        return date_type.fromisoformat(str(value).strip())
    except ValueError:
        raise HTTPException(400, f"Date '{value}' must be YYYY-MM-DD.")


def _tests_dict(lot: models.TestSampleLot):
    """Lot.tests is stored as a JSON string; never let bad data 500 the page."""
    if not lot.tests:
        return {}
    try:
        val = json.loads(lot.tests)
        return val if isinstance(val, dict) else {}
    except (ValueError, TypeError):
        return {}


def _slugify(label: str) -> str:
    out = "".join(c.lower() if c.isalnum() else "_" for c in label)
    while "__" in out:
        out = out.replace("__", "_")
    return out.strip("_")[:64] or "test"


def _criteria(db: Session, include_inactive=False):
    q = db.query(models.TestCriterion)
    if not include_inactive:
        q = q.filter(models.TestCriterion.active == True)
    return q.order_by(models.TestCriterion.sort_order, models.TestCriterion.id).all()


def _lot_brief(lot: models.TestSampleLot, minis=None):
    return {
        "id": lot.id, "campaign_id": lot.campaign_id, "lot_id": lot.lot_id,
        "build": lot.build, "requestor": lot.requestor,
        "completion_date": lot.completion_date.isoformat() if lot.completion_date else None,
        "location": lot.location, "archive_location": lot.archive_location,
        "comments": lot.comments, "tests": _tests_dict(lot),
        "checked_out": bool(lot.checked_out), "held_by": lot.held_by,
        "checked_out_at": lot.checked_out_at.isoformat(timespec="seconds") if lot.checked_out_at else None,
        "minis": [{"id": m.id, "mini_id": m.mini_id, "location": m.location,
                    "note": m.note, "tests": _tests_dict(m)}
                  for m in (minis if minis is not None else lot.minis)],
    }


# ─────────────────────────────────────────────────────────── the whole tree
@router.get("/tree")
def testing_tree(db: Session = Depends(get_db), user=Depends(current_user)):
    """One call returns everything the view needs: departments with their
    campaigns, lots and minis, plus the active criteria that become columns."""
    crits = _criteria(db)
    campaigns = (db.query(models.TestCampaign)
                 .filter(models.TestCampaign.active == True)
                 .order_by(models.TestCampaign.name).all())

    # Preload lots and minis so this isn't N+1 across hundreds of rows.
    camp_ids = [c.id for c in campaigns]
    lots_by_camp = {}
    minis_by_lot = {}
    if camp_ids:
        lots = (db.query(models.TestSampleLot)
                .filter(models.TestSampleLot.campaign_id.in_(camp_ids))
                .order_by(models.TestSampleLot.lot_id).all())
        lot_ids = [l.id for l in lots]
        if lot_ids:
            for m in (db.query(models.TestSampleMini)
                      .filter(models.TestSampleMini.lot_id.in_(lot_ids))
                      .order_by(models.TestSampleMini.id).all()):
                minis_by_lot.setdefault(m.lot_id, []).append(m)
        for l in lots:
            lots_by_camp.setdefault(l.campaign_id, []).append(l)

    depts = {d.id: d.name for d in db.query(models.Department).all()}
    groups = {}
    for c in campaigns:
        key = c.department_id
        groups.setdefault(key, []).append({
            "id": c.id, "name": c.name, "description": c.description,
            "department_id": c.department_id,
            "lots": [_lot_brief(l, minis_by_lot.get(l.id, []))
                     for l in lots_by_camp.get(c.id, [])],
        })

    tree = []
    for dept_id, camps in groups.items():
        tree.append({
            "department_id": dept_id,
            "department": depts.get(dept_id) or "(no department)",
            "campaigns": camps,
        })
    # named departments first, "(no department)" last
    tree.sort(key=lambda g: (g["department_id"] is None, (g["department"] or "").lower()))

    return {
        "tree": tree,
        "criteria": [{"id": c.id, "key": c.key, "label": c.label} for c in crits],
        "states": STATES,
    }


# ─────────────────────────────────────────────────────────── campaigns
@router.post("/campaigns")
def create_campaign(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Campaign name is required.")
    dept_id = body.get("department_id") or None
    if dept_id and not db.get(models.Department, int(dept_id)):
        raise HTTPException(400, "Unknown department.")
    c = models.TestCampaign(name=name,
                            department_id=int(dept_id) if dept_id else None,
                            description=(body.get("description") or "").strip() or None)
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "name": c.name, "department_id": c.department_id}


@router.put("/campaigns/{campaign_id}")
def update_campaign(campaign_id: int, body: dict, db: Session = Depends(get_db),
                    user=Depends(current_user)):
    c = db.get(models.TestCampaign, campaign_id)
    if not c:
        raise HTTPException(404, "Campaign not found")
    if "name" in body:
        name = (body.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "Campaign name is required.")
        c.name = name
    if "department_id" in body:
        dept_id = body.get("department_id") or None
        if dept_id and not db.get(models.Department, int(dept_id)):
            raise HTTPException(400, "Unknown department.")
        c.department_id = int(dept_id) if dept_id else None
    if "description" in body:
        c.description = (body.get("description") or "").strip() or None
    db.commit(); db.refresh(c)
    return {"id": c.id, "name": c.name, "department_id": c.department_id}


@router.delete("/campaigns/{campaign_id}")
def delete_campaign(campaign_id: int, db: Session = Depends(get_db), user=Depends(current_user)):
    c = db.get(models.TestCampaign, campaign_id)
    if not c:
        raise HTTPException(404, "Campaign not found")
    db.delete(c); db.commit()          # cascades to lots -> minis
    return {"deleted": True}


# ─────────────────────────────────────────────────────────── lots
@router.post("/lots")
def create_lot(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    campaign_id = body.get("campaign_id")
    if not campaign_id or not db.get(models.TestCampaign, int(campaign_id)):
        raise HTTPException(400, "A valid campaign is required.")
    lot_id = (body.get("lot_id") or "").strip()
    if not lot_id:
        raise HTTPException(400, "Lot ID is required.")
    l = models.TestSampleLot(
        campaign_id=int(campaign_id), lot_id=lot_id,
        build=(body.get("build") or "").strip() or None,
        requestor=(body.get("requestor") or "").strip() or None,
        completion_date=_parse_date(body.get("completion_date")),
        location=(body.get("location") or "").strip() or None,
        archive_location=(body.get("archive_location") or "").strip() or None,
        comments=(body.get("comments") or "").strip() or None)
    db.add(l); db.commit(); db.refresh(l)
    return _lot_brief(l, [])


@router.put("/lots/{lot_pk}")
def update_lot(lot_pk: int, body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    l = db.get(models.TestSampleLot, lot_pk)
    if not l:
        raise HTTPException(404, "Lot not found")
    if "lot_id" in body:
        v = (body.get("lot_id") or "").strip()
        if not v:
            raise HTTPException(400, "Lot ID is required.")
        l.lot_id = v
    for field in ("build", "requestor", "location", "archive_location", "comments"):
        if field in body:
            setattr(l, field, (body.get(field) or "").strip() or None)
    if "completion_date" in body:
        l.completion_date = _parse_date(body.get("completion_date"))
    # Setting the location on a lot pushes it down to every mini under it —
    # minis physically travel with their lot, so keeping them in sync by hand
    # was pure busywork. Pass cascade=false to edit the lot alone.
    if "location" in body and body.get("cascade", True):
        for m in l.minis:
            m.location = l.location
    if "campaign_id" in body and body["campaign_id"]:
        if not db.get(models.TestCampaign, int(body["campaign_id"])):
            raise HTTPException(400, "Unknown campaign.")
        l.campaign_id = int(body["campaign_id"])
    db.commit(); db.refresh(l)
    return _lot_brief(l)


@router.put("/lots/{lot_pk}/tests")
def set_lot_test(lot_pk: int, body: dict, db: Session = Depends(get_db),
                 user=Depends(current_user)):
    """Set one criterion's status on a lot. Writes into the JSON blob rather
    than a column, which is what makes adding a test column migration-free."""
    l = db.get(models.TestSampleLot, lot_pk)
    if not l:
        raise HTTPException(404, "Lot not found")
    key = (body.get("key") or "").strip()
    state = (body.get("state") or "").strip()
    if not key:
        raise HTTPException(400, "A criterion key is required.")
    if state not in STATES:
        raise HTTPException(400, f"State must be one of: {', '.join(s or '(none)' for s in STATES)}")
    valid = {c.key for c in _criteria(db, include_inactive=True)}
    if key not in valid:
        raise HTTPException(400, "Unknown test criterion.")
    data = _tests_dict(l)
    if state:
        data[key] = state
    else:
        data.pop(key, None)          # "not started" is absence, not a value
    l.tests = json.dumps(data)
    db.commit()
    return {"id": l.id, "tests": data}


@router.delete("/lots/{lot_pk}")
def delete_lot(lot_pk: int, db: Session = Depends(get_db), user=Depends(current_user)):
    l = db.get(models.TestSampleLot, lot_pk)
    if not l:
        raise HTTPException(404, "Lot not found")
    db.delete(l); db.commit()        # cascades to minis
    return {"deleted": True}


# ─────────────────────────────────────────────────────────── minis
@router.post("/minis")
def create_mini(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    lot_pk = body.get("lot_id")
    if not lot_pk or not db.get(models.TestSampleLot, int(lot_pk)):
        raise HTTPException(400, "A valid lot is required.")
    mini_id = (body.get("mini_id") or "").strip()
    if not mini_id:
        raise HTTPException(400, "Mini ID is required.")
    m = models.TestSampleMini(lot_id=int(lot_pk), mini_id=mini_id,
                              location=(body.get("location") or "").strip() or None,
                              note=(body.get("note") or "").strip() or None)
    db.add(m); db.commit(); db.refresh(m)
    return {"id": m.id, "mini_id": m.mini_id, "location": m.location,
            "note": m.note, "lot_id": m.lot_id}


@router.post("/minis/bulk")
def create_minis_bulk(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    """Paste many mini IDs at once, one per line (or comma-separated). Lots in
    the source workbook routinely hold a dozen or more minis, so adding them
    one at a time is the slowest part of the whole workflow."""
    lot_pk = body.get("lot_id")
    if not lot_pk or not db.get(models.TestSampleLot, int(lot_pk)):
        raise HTTPException(400, "A valid lot is required.")
    ids = [t.strip() for t in re.split(r"[\n,]+", str(body.get("mini_ids") or "")) if t.strip()]
    if not ids:
        raise HTTPException(400, "No mini IDs provided.")
    existing = {m.mini_id for m in
                db.query(models.TestSampleMini)
                  .filter(models.TestSampleMini.lot_id == int(lot_pk)).all()}
    made = 0
    for mid in ids:
        if mid in existing:          # don't duplicate on a re-paste
            continue
        db.add(models.TestSampleMini(lot_id=int(lot_pk), mini_id=mid))
        existing.add(mid); made += 1
    db.commit()
    return {"created": made, "skipped_duplicates": len(ids) - made}


@router.put("/minis/{mini_pk}")
def update_mini(mini_pk: int, body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    m = db.get(models.TestSampleMini, mini_pk)
    if not m:
        raise HTTPException(404, "Mini not found")
    if "mini_id" in body:
        v = (body.get("mini_id") or "").strip()
        if not v:
            raise HTTPException(400, "Mini ID is required.")
        m.mini_id = v
    if "location" in body:
        m.location = (body.get("location") or "").strip() or None
    if "note" in body:
        m.note = (body.get("note") or "").strip() or None
    db.commit(); db.refresh(m)
    return {"id": m.id, "mini_id": m.mini_id, "location": m.location,
            "note": m.note, "lot_id": m.lot_id}


@router.put("/minis/{mini_pk}/tests")
def set_mini_test(mini_pk: int, body: dict, db: Session = Depends(get_db),
                  user=Depends(current_user)):
    """Set one criterion on one mini. Results are per-mini because an
    individual mini can fail a test the rest of its lot passed."""
    m = db.get(models.TestSampleMini, mini_pk)
    if not m:
        raise HTTPException(404, "Mini not found")
    key = (body.get("key") or "").strip()
    state = (body.get("state") or "").strip()
    if not key:
        raise HTTPException(400, "A criterion key is required.")
    if state not in STATES:
        raise HTTPException(400, "State must be blank or 'done'.")
    valid = {c.key for c in _criteria(db, include_inactive=True)}
    if key not in valid:
        raise HTTPException(400, "Unknown test criterion.")
    data = _tests_dict(m)
    if state:
        data[key] = state
    else:
        data.pop(key, None)
    m.tests = json.dumps(data)
    db.commit()
    return {"id": m.id, "tests": data}


@router.put("/lots/{lot_pk}/tests-all")
def set_lot_test_all(lot_pk: int, body: dict, db: Session = Depends(get_db),
                     user=Depends(current_user)):
    """Set one criterion across every mini in a lot — the common case when a
    whole lot passes together."""
    lot = db.get(models.TestSampleLot, lot_pk)
    if not lot:
        raise HTTPException(404, "Lot not found")
    key = (body.get("key") or "").strip()
    state = (body.get("state") or "").strip()
    if state not in STATES:
        raise HTTPException(400, "State must be blank or 'done'.")
    valid = {c.key for c in _criteria(db, include_inactive=True)}
    if key not in valid:
        raise HTTPException(400, "Unknown test criterion.")
    for m in lot.minis:
        data = _tests_dict(m)
        if state:
            data[key] = state
        else:
            data.pop(key, None)
        m.tests = json.dumps(data)
    db.commit()
    return {"lot_id": lot.id, "minis": len(lot.minis), "key": key, "state": state}


@router.delete("/minis/{mini_pk}")
def delete_mini(mini_pk: int, db: Session = Depends(get_db), user=Depends(current_user)):
    m = db.get(models.TestSampleMini, mini_pk)
    if not m:
        raise HTTPException(404, "Mini not found")
    db.delete(m); db.commit()
    return {"deleted": True}


# ─────────────────────────────────────────── criteria (the columns) - admin
@router.get("/criteria")
def list_criteria(db: Session = Depends(get_db), user=Depends(current_user)):
    return [{"id": c.id, "key": c.key, "label": c.label,
             "sort_order": c.sort_order, "active": c.active}
            for c in _criteria(db, include_inactive=True)]


@router.post("/criteria")
def create_criterion(body: dict, db: Session = Depends(get_db), user=Depends(require_admin)):
    label = (body.get("label") or "").strip()
    if not label:
        raise HTTPException(400, "Label is required.")
    key = (body.get("key") or "").strip() or _slugify(label)
    existing = db.query(models.TestCriterion).filter(models.TestCriterion.key == key).first()
    if existing:
        # Re-activating a soft-deleted column brings its old values back with it.
        if not existing.active:
            existing.active = True
            existing.label = label
            db.commit(); db.refresh(existing)
            return {"id": existing.id, "key": existing.key, "label": existing.label,
                    "reactivated": True}
        raise HTTPException(400, "A test column with that key already exists.")
    top = db.query(func.max(models.TestCriterion.sort_order)).scalar() or 0
    c = models.TestCriterion(key=key, label=label, sort_order=top + 1, active=True)
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "key": c.key, "label": c.label}


@router.put("/criteria/{crit_id}")
def update_criterion(crit_id: int, body: dict, db: Session = Depends(get_db),
                     user=Depends(require_admin)):
    c = db.get(models.TestCriterion, crit_id)
    if not c:
        raise HTTPException(404, "Test column not found")
    if "label" in body:
        label = (body.get("label") or "").strip()
        if not label:
            raise HTTPException(400, "Label is required.")
        c.label = label            # key deliberately never changes - lots reference it
    if "sort_order" in body:
        c.sort_order = int(body["sort_order"])
    if "active" in body:
        c.active = bool(body["active"])
    db.commit(); db.refresh(c)
    return {"id": c.id, "key": c.key, "label": c.label,
            "sort_order": c.sort_order, "active": c.active}


@router.delete("/criteria/{crit_id}")
def delete_criterion(crit_id: int, db: Session = Depends(get_db), user=Depends(require_admin)):
    """Soft delete. The column vanishes from the grid but any statuses already
    recorded against it survive in each lot's JSON, so re-adding it restores
    the history instead of losing it."""
    c = db.get(models.TestCriterion, crit_id)
    if not c:
        raise HTTPException(404, "Test column not found")
    c.active = False
    db.commit()
    return {"deleted": True, "soft": True}


# ------------------------------------------------------------------ custody
# Test samples are lent to a named person rather than a department, so custody
# here is a free-text holder — deliberately different from the asset dashboard.
@router.post("/lots/{lot_pk}/checkout")
def checkout_lot(lot_pk: int, body: dict, db: Session = Depends(get_db),
                 user=Depends(current_user)):
    lot = db.get(models.TestSampleLot, lot_pk)
    if not lot:
        raise HTTPException(404, "Lot not found")
    held_by = (body.get("held_by") or "").strip()
    if not held_by:
        raise HTTPException(400, "A name is required to check out.")
    lot.checked_out = True
    lot.held_by = held_by
    lot.checked_out_at = models.utcnow()
    db.commit(); db.refresh(lot)
    return _lot_brief(lot)


@router.post("/lots/{lot_pk}/return")
def return_lot(lot_pk: int, db: Session = Depends(get_db), user=Depends(current_user)):
    lot = db.get(models.TestSampleLot, lot_pk)
    if not lot:
        raise HTTPException(404, "Lot not found")
    lot.checked_out = False
    lot.held_by = None
    lot.checked_out_at = None
    db.commit(); db.refresh(lot)
    return _lot_brief(lot)


@router.post("/scan")
def testing_scan(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    """Barcode/typed scan against a Lot ID or Mini ID.

    modes: lookup | checkout | return
    A mini scan resolves to its parent lot, since custody is tracked per lot.
    """
    code = (body.get("code") or "").strip()
    mode = (body.get("mode") or "lookup").strip()
    if not code:
        raise HTTPException(400, "No code scanned.")

    lot = (db.query(models.TestSampleLot)
           .filter(func.lower(models.TestSampleLot.lot_id) == code.lower()).first())
    matched_mini = None
    if not lot:
        matched_mini = (db.query(models.TestSampleMini)
                        .filter(func.lower(models.TestSampleMini.mini_id) == code.lower()).first())
        if matched_mini:
            lot = db.get(models.TestSampleLot, matched_mini.lot_id)
    if not lot:
        raise HTTPException(404, f"No lot or mini matches '{code}'.")

    if mode == "checkout":
        held_by = (body.get("held_by") or "").strip()
        if not held_by:
            raise HTTPException(400, "Enter a name before scanning.")
        lot.checked_out = True
        lot.held_by = held_by
        lot.checked_out_at = models.utcnow()
        message = f"checked out to {held_by}"
    elif mode == "return":
        if not lot.checked_out:
            message = "was already available"
        else:
            message = f"returned from {lot.held_by}"
        lot.checked_out = False
        lot.held_by = None
        lot.checked_out_at = None
    else:
        message = ("checked out to " + lot.held_by) if lot.checked_out else "available"
    db.commit(); db.refresh(lot)

    return {"ok": True, "message": message, "mode": mode,
            "matched": "mini" if matched_mini else "lot",
            "matched_id": matched_mini.mini_id if matched_mini else lot.lot_id,
            "mini_id": matched_mini.id if matched_mini else None,
            "lot": _lot_brief(lot)}


# ------------------------------------------------------------------ bulk edit
@router.post("/bulk-edit")
def testing_bulk_edit(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    """Apply the same field values to many lots and/or minis at once.

    Selecting a lot implies its minis (the frontend sends both), so editing a
    location across a whole campaign is one action rather than dozens.
    Only keys present in `fields` are touched; an explicit "" clears.
    """
    lot_ids = [int(i) for i in (body.get("lot_ids") or [])]
    mini_ids = [int(i) for i in (body.get("mini_ids") or [])]
    fields = body.get("fields") or {}
    if not lot_ids and not mini_ids:
        raise HTTPException(400, "Nothing selected.")
    if not fields:
        raise HTTPException(400, "Nothing to change.")

    lot_allowed = {"build", "requestor", "location", "archive_location",
                   "comments", "completion_date"}
    mini_allowed = {"location", "note"}
    unknown = set(fields) - (lot_allowed | mini_allowed)
    if unknown:
        raise HTTPException(400, f"Can't bulk-edit: {', '.join(sorted(unknown))}")

    touched_lots = touched_minis = 0
    if lot_ids:
        for lot in db.query(models.TestSampleLot).filter(
                models.TestSampleLot.id.in_(lot_ids)).all():
            for k, v in fields.items():
                if k not in lot_allowed:
                    continue
                if k == "completion_date":
                    lot.completion_date = _parse_date(v)
                else:
                    setattr(lot, k, (str(v).strip() or None) if v is not None else None)
            # keep minis aligned with their lot's location
            if "location" in fields:
                for m in lot.minis:
                    m.location = lot.location
            touched_lots += 1
    if mini_ids:
        for m in db.query(models.TestSampleMini).filter(
                models.TestSampleMini.id.in_(mini_ids)).all():
            for k, v in fields.items():
                if k not in mini_allowed:
                    continue
                setattr(m, k, (str(v).strip() or None) if v is not None else None)
            touched_minis += 1
    db.commit()
    return {"lots": touched_lots, "minis": touched_minis}


# ------------------------------------------------------------------ mini detail
@router.get("/minis/{mini_pk}")
def get_mini(mini_pk: int, db: Session = Depends(get_db), user=Depends(current_user)):
    """Everything about one mini, including the lot and campaign it sits under
    so the detail view can show inherited context."""
    m = db.get(models.TestSampleMini, mini_pk)
    if not m:
        raise HTTPException(404, "Mini not found")
    lot = db.get(models.TestSampleLot, m.lot_id)
    camp = db.get(models.TestCampaign, lot.campaign_id) if lot else None
    dept = db.get(models.Department, camp.department_id) if camp and camp.department_id else None
    crits = _criteria(db)
    tests = _tests_dict(lot) if lot else {}
    return {
        "id": m.id, "mini_id": m.mini_id, "location": m.location, "note": m.note,
        "created_at": m.created_at.isoformat(timespec="seconds") if m.created_at else None,
        "lot": _lot_brief(lot) if lot else None,
        "campaign": {"id": camp.id, "name": camp.name} if camp else None,
        "department": dept.name if dept else None,
        "criteria": [{"key": c.key, "label": c.label,
                      "state": _tests_dict(m).get(c.key, "")} for c in crits],
    }


# ------------------------------------------------------------------ excel import
# The source sheets put one Lot ID against many Mini rows, with the lot-level
# values (description, date, requestor) filled only on the first row of the
# group and blank on the rest. The importer walks rows in order and carries
# those values forward, which is what turns a flat sheet back into a hierarchy.
TESTING_IMPORT_FIELDS = ["lot_id", "mini_id", "build", "completion_date",
                         "requestor", "location", "archive_location", "comments"]


def _cellval(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    return v


@router.post("/import/inspect")
async def testing_import_inspect(file: UploadFile = File(...), header_row: int = Form(1),
                                 sheet: str = Form(""), user=Depends(current_user)):
    """Return the sheet names and column headers so the UI can build the
    mapping dropdowns (including the manual Field #1..#N test columns)."""
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Couldn't read that file ({e}).")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers, rows = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == header_row:
            headers = [str(c).strip() if c is not None else "" for c in row]
        elif i > header_row:
            rows.append(row)
            if len(rows) >= 8:
                break
    wb.close()
    return {"sheets": wb.sheetnames if hasattr(wb, "sheetnames") else [ws.title],
            "sheet": ws.title,
            "columns": [h for h in headers if h],
            "all_columns": headers,
            "sample_rows": [[("" if c is None else str(c)[:40]) for c in r] for r in rows],
            "fields": TESTING_IMPORT_FIELDS}


def _read_testing_rows(raw, sheet, header_row, mapping, test_fields):
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers, data = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == header_row:
            headers = [str(c).strip() if c is not None else "" for c in row]
        elif i > header_row:
            data.append(row)
    wb.close()
    idx = {h: i for i, h in enumerate(headers) if h}
    for field, col in list(mapping.items()) + [(t["key"], t["column"]) for t in test_fields]:
        if col and col not in idx:
            raise HTTPException(400, f"Column '{col}' isn't in the header row.")

    def cell(row, col):
        if not col or col not in idx or idx[col] >= len(row):
            return None
        return _cellval(row[idx[col]])

    return data, cell


@router.post("/import/commit")
async def testing_import_commit(file: UploadFile = File(...), header_row: int = Form(1),
                                sheet: str = Form(""), mapping: str = Form(...),
                                test_fields: str = Form("[]"),
                                campaign_id: int = Form(...),
                                db: Session = Depends(get_db), user=Depends(current_user)):
    """Import a sheet into ONE campaign (which already carries the department).

    `mapping` maps our field names to column headers.
    `test_fields` is a list of {label, column} — the manually chosen Field #1,
    Field #2... Each becomes a TestCriterion if it doesn't exist, and a truthy
    cell records a check against the lot.
    """
    camp = db.get(models.TestCampaign, campaign_id)
    if not camp:
        raise HTTPException(400, "Pick a test campaign to import into.")
    m = json.loads(mapping)
    tf = json.loads(test_fields)
    if not m.get("lot_id"):
        raise HTTPException(400, "Map a column to Lot ID.")

    # Create (or reuse) a criterion per chosen test field.
    resolved = []
    for item in tf:
        label = (item.get("label") or "").strip()
        col = (item.get("column") or "").strip()
        if not label or not col:
            continue
        key = _slugify(label)
        crit = db.query(models.TestCriterion).filter(models.TestCriterion.key == key).first()
        if not crit:
            top = db.query(func.max(models.TestCriterion.sort_order)).scalar() or 0
            crit = models.TestCriterion(key=key, label=label, sort_order=top + 1, active=True)
            db.add(crit); db.flush()
        elif not crit.active:
            crit.active = True
        resolved.append({"key": key, "column": col})

    raw = await file.read()
    data, cell = _read_testing_rows(raw, sheet, header_row, m, resolved)

    existing_lots = {l.lot_id: l for l in
                     db.query(models.TestSampleLot)
                       .filter(models.TestSampleLot.campaign_id == camp.id).all()}
    lots_made = minis_made = skipped = 0
    current = None          # the lot the following mini-only rows belong to

    for row in data:
        lot_code = cell(row, m.get("lot_id"))
        mini_code = cell(row, m.get("mini_id"))
        lot_code = str(lot_code).strip() if lot_code is not None else None
        mini_code = str(mini_code).strip() if mini_code is not None else None

        # A row with a lot value but no mini is a week/section banner, not data.
        if lot_code and not mini_code:
            current = None
            skipped += 1
            continue
        if not lot_code and not mini_code:
            skipped += 1
            continue

        if lot_code:
            lot = existing_lots.get(lot_code)
            if not lot:
                lot = models.TestSampleLot(
                    campaign_id=camp.id, lot_id=lot_code,
                    build=_as_text(cell(row, m.get("build"))),
                    requestor=_as_text(cell(row, m.get("requestor"))),
                    completion_date=_parse_date_loose(cell(row, m.get("completion_date"))),
                    location=_as_text(cell(row, m.get("location"))),
                    archive_location=_as_text(cell(row, m.get("archive_location"))),
                    comments=_as_text(cell(row, m.get("comments"))),
                    tests=json.dumps({}))
                db.add(lot); db.flush()
                existing_lots[lot_code] = lot
                lots_made += 1
            current = lot

        if mini_code and current is not None:
            exists = (db.query(models.TestSampleMini)
                      .filter(models.TestSampleMini.lot_id == current.id,
                              models.TestSampleMini.mini_id == mini_code).first())
            # Each mini row carries its own TRUE/FALSE cells, so results are
            # read per row rather than once per lot.
            row_tests = {}
            for r in resolved:
                if _truthy(cell(row, r["column"])):
                    row_tests[r["key"]] = "done"
            if exists:
                if row_tests:
                    merged = _tests_dict(exists); merged.update(row_tests)
                    exists.tests = json.dumps(merged)
            else:
                db.add(models.TestSampleMini(lot_id=current.id, mini_id=mini_code,
                                             location=current.location,
                                             tests=json.dumps(row_tests)))
                minis_made += 1

    db.commit()
    return {"campaign": camp.name, "lots_created": lots_made,
            "minis_created": minis_made, "rows_skipped": skipped}
