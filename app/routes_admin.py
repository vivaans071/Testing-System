"""
Admin-facing API. Reads require any signed-in user (admin or viewer);
writes require admin. QR image / label routes stay open (see comments below)
since browsers can't attach an Authorization header to <img src> or
window.open(), and both only ever lead to the already-public check-in page.
"""
import datetime
import io
import json
import re
from html import escape as _esc
from datetime import date as date_type, datetime as _dt
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session

from . import models, notify
from .barcode_gen import make_barcode_png
from .database import get_db
from .qr import checkin_url, make_qr_png
from .security import current_user, require_admin, require_support, require_owner

router = APIRouter(prefix="/api", tags=["admin"])

DEFAULT_BASE_URL = "http://localhost:8000"
DEFAULT_COMPANY_NAME = "Frore Systems"


def get_setting(db: Session, key: str, default: str) -> str:
    row = db.get(models.Setting, key)
    return row.value if row and row.value else default


def base_url(db: Session) -> str:
    return get_setting(db, "base_url", DEFAULT_BASE_URL)


def company_name(db: Session) -> str:
    return get_setting(db, "company_name", DEFAULT_COMPANY_NAME)


def qualified_emails(db: Session) -> list:
    """Comma-separated recipients for the 'asset became Qualified' notice."""
    raw = get_setting(db, "qualified_notify_emails", "")
    return [e.strip() for e in raw.split(",") if e.strip()]


def notify_qualified(db: Session, a, changed_by, note):
    """Fire the Qualified email. Never raises - a mail problem must not break
    a status change."""
    try:
        to = qualified_emails(db)
        if not to or not notify.is_configured():
            return
        url = f"{base_url(db)}/c/{a.code}"
        subject, body = notify.qualified_email(a, changed_by, note, url)
        notify.send_async(to, subject, body)
    except Exception as e:                                  # noqa: BLE001
        print(f"[notify] skipped: {e}")


def _parse_cost(value):
    """Accepts numbers or strings like '$1,200.50'. Returns None for blank,
    raises HTTPException(400) for genuinely unparseable input instead of
    letting a raw ValueError become a 500."""
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace("$", "").replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        raise HTTPException(400, f"Cost '{value}' isn't a valid number.")


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, date_type):
        return value
    try:
        return date_type.fromisoformat(str(value).strip())
    except ValueError:
        raise HTTPException(400, f"Date '{value}' must be YYYY-MM-DD (the form sends this automatically).")


# --------------------------------------------------------------- settings
@router.get("/settings")
def read_settings(db: Session = Depends(get_db), user=Depends(current_user)):
    return {"base_url": base_url(db), "company_name": company_name(db),
            "statuses": models.STATUSES,
            "qualified_notify_emails": get_setting(db, "qualified_notify_emails", ""),
            "email_configured": notify.is_configured()}


@router.put("/settings")
def write_settings(body: dict, db: Session = Depends(get_db), user=Depends(require_admin)):
    if "base_url" in body:
        row = db.get(models.Setting, "base_url") or models.Setting(key="base_url")
        row.value = body["base_url"].rstrip("/")
        db.merge(row)
        db.commit()
    if "company_name" in body:
        row = db.get(models.Setting, "company_name") or models.Setting(key="company_name")
        row.value = body["company_name"].strip()
        db.merge(row)
        db.commit()
    if "qualified_notify_emails" in body:
        cleaned = ", ".join(e.strip() for e in (body["qualified_notify_emails"] or "").split(",")
                            if e.strip())
        row = (db.get(models.Setting, "qualified_notify_emails")
               or models.Setting(key="qualified_notify_emails"))
        row.value = cleaned
        db.merge(row)
        db.commit()
    return read_settings(db)


# --------------------------------------------------------------- departments
# Known Frore department codes, used to prefill the code when a department is
# created with a recognisable name. Anything else the admin types in by hand.
DEPT_CODE_HINTS = {
    "general/administration": "ADMN", "general / administration": "ADMN",
    "administration": "ADMN", "general": "ADMN",
    "airjet design, acoustic, thermal": "DETC", "airjet design": "DETC",
    "asic": "ASIC", "data center": "DATA", "datacenter": "DATA",
    "process technology": "TECH", "rd corporate": "CORP", "r&d corporate": "CORP",
    "software, firmware, sq&a": "SQSA", "software": "SQSA",
    "ce": "CENG", "business development": "BSDV", "marketing": "MKTG",
}


def _guess_dept_code(name: str) -> str:
    return DEPT_CODE_HINTS.get((name or "").strip().lower(), "")


@router.get("/departments")
def list_departments(db: Session = Depends(get_db), user=Depends(current_user)):
    depts = db.query(models.Department).order_by(models.Department.name).all()
    counts = {d.id: 0 for d in depts}
    for a in db.query(models.Asset).filter(models.Asset.department_id.isnot(None)):
        counts[a.department_id] = counts.get(a.department_id, 0) + 1
    return [{"id": d.id, "name": d.name, "code": d.code or "",
             "asset_count": counts.get(d.id, 0)} for d in depts]


@router.post("/departments")
def create_department(body: dict, db: Session = Depends(get_db), user=Depends(require_admin)):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Name is required")
    if db.query(models.Department).filter(models.Department.name == name).first():
        raise HTTPException(400, "That department already exists")
    code = (body.get("code") or "").strip().upper() or _guess_dept_code(name)
    d = models.Department(name=name, code=code or None)
    db.add(d); db.commit(); db.refresh(d)
    return {"id": d.id, "name": d.name, "code": d.code or "", "asset_count": 0}


@router.put("/departments/{dept_id}")
def update_department(dept_id: int, body: dict, db: Session = Depends(get_db),
                      user=Depends(require_admin)):
    d = db.get(models.Department, dept_id)
    if not d:
        raise HTTPException(404, "Department not found")
    if "name" in body:
        name = (body.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "Name is required")
        clash = db.query(models.Department).filter(models.Department.name == name,
                                                   models.Department.id != dept_id).first()
        if clash:
            raise HTTPException(400, "That department already exists")
        d.name = name
    if "code" in body:
        # Codes already baked into printed asset codes don't change retroactively;
        # editing this only affects codes generated from here on.
        d.code = (body.get("code") or "").strip().upper() or None
    db.commit(); db.refresh(d)
    return {"id": d.id, "name": d.name, "code": d.code or ""}


@router.delete("/departments/{dept_id}")
def delete_department(dept_id: int, db: Session = Depends(get_db), user=Depends(require_admin)):
    d = db.get(models.Department, dept_id)
    if not d:
        raise HTTPException(404, "Department not found")
    db.query(models.Asset).filter(models.Asset.department_id == dept_id) \
        .update({"department_id": None})
    db.delete(d); db.commit()
    return {"deleted": True}



# --------------------------------------------------------------- owners
OWNER_SETTING_KEY = "owners_json"


def _owner_list(db: Session):
    row = db.get(models.Setting, OWNER_SETTING_KEY)
    if not row or not row.value:
        return []
    try:
        data = json.loads(row.value)
        return sorted([str(x).strip() for x in data if str(x).strip()], key=str.lower)
    except Exception:
        return []


def _save_owner_list(db: Session, owners):
    clean = sorted(set(o.strip() for o in owners if o and o.strip()), key=str.lower)
    row = db.get(models.Setting, OWNER_SETTING_KEY) or models.Setting(key=OWNER_SETTING_KEY)
    row.value = json.dumps(clean)
    db.merge(row)
    db.commit()
    return clean


@router.get("/owners")
def list_owners(db: Session = Depends(get_db), user=Depends(current_user)):
    owners = _owner_list(db)
    counts = {o: 0 for o in owners}
    for a in db.query(models.Asset).filter(models.Asset.owner.isnot(None)):
        if a.owner:
            counts[a.owner] = counts.get(a.owner, 0) + 1
    return [{"name": o, "asset_count": counts.get(o, 0)} for o in sorted(counts, key=str.lower)]


@router.post("/owners")
def create_owner(body: dict, db: Session = Depends(get_db), user=Depends(require_owner)):
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Name is required")
    owners = _owner_list(db)
    if name.lower() not in [o.lower() for o in owners]:
        owners.append(name)
    _save_owner_list(db, owners)
    return {"name": name}


@router.delete("/owners/{owner_name}")
def delete_owner(owner_name: str, db: Session = Depends(get_db), user=Depends(require_owner)):
    owners = [o for o in _owner_list(db) if o.lower() != owner_name.lower()]
    _save_owner_list(db, owners)
    return {"deleted": True}

# --------------------------------------------------------------- activity feed
@router.get("/activity")
def activity_feed(limit: int = 50, before: str = "", db: Session = Depends(get_db),
                  user=Depends(current_user)):
    """Merged, newest-first feed of the two things worth noticing across all
    assets: notes left on check-ins, and real status transitions. Bare check-ins
    (no note) are intentionally excluded to keep the feed signal-heavy.

    Pass before=<ts of the oldest row you have> to page backwards ('load more').
    has_more tells the UI whether an older page exists."""
    limit = max(1, min(limit, 200))
    fetch = limit + 1                      # one extra row tells us if more exist
    events = []

    note_q = (db.query(models.CheckIn)
              .filter(models.CheckIn.note.isnot(None), models.CheckIn.note != ""))
    change_q = (db.query(models.StatusChange)
                .filter(models.StatusChange.old_status.isnot(None),
                        models.StatusChange.old_status != models.StatusChange.new_status))
    if before:
        try:
            cutoff = _dt.fromisoformat(before.replace("Z", ""))
            note_q = note_q.filter(models.CheckIn.ts < cutoff)
            change_q = change_q.filter(models.StatusChange.ts < cutoff)
        except ValueError:
            pass

    for c in note_q.order_by(models.CheckIn.ts.desc()).limit(fetch).all():
        a = c.asset
        events.append({
            "type": "note", "ts": c.ts.isoformat(), "asset_id": c.asset_id,
            "asset_name": a.name if a else "(deleted asset)",
            "asset_code": a.code if a else None,
            "asset_active": bool(a.active) if a else False,
            "who": c.checked_in_by or None, "note": c.note, "source": c.source,
        })
    for s in change_q.order_by(models.StatusChange.ts.desc()).limit(fetch).all():
        a = s.asset
        events.append({
            "type": "status", "ts": s.ts.isoformat(), "asset_id": s.asset_id,
            "asset_name": a.name if a else "(deleted asset)",
            "asset_code": a.code if a else None,
            "asset_active": bool(a.active) if a else False,
            "who": s.changed_by or None, "note": s.note,
            "old_status": s.old_status, "new_status": s.new_status,
        })

    events.sort(key=lambda e: e["ts"], reverse=True)
    has_more = len(events) > limit
    return {"events": events[:limit], "has_more": has_more,
            "oldest": events[:limit][-1]["ts"] if events[:limit] else None}


@router.get("/activity/unread")
def activity_unread(since: str = "", db: Session = Depends(get_db),
                    user=Depends(current_user)):
    """Cheap count-only companion to /activity for the notification badge.
    Returns how many feed events are newer than `since`, plus the newest ts."""
    note_f = (models.CheckIn.note.isnot(None), models.CheckIn.note != "")
    stat_f = (models.StatusChange.old_status.isnot(None),
              models.StatusChange.old_status != models.StatusChange.new_status)

    q_notes = db.query(func.count(models.CheckIn.id)).filter(*note_f)
    q_stat = db.query(func.count(models.StatusChange.id)).filter(*stat_f)
    if since:
        try:
            dt = _dt.fromisoformat(since.replace("Z", ""))
            q_notes = q_notes.filter(models.CheckIn.ts > dt)
            q_stat = q_stat.filter(models.StatusChange.ts > dt)
        except ValueError:
            pass
    count = (q_notes.scalar() or 0) + (q_stat.scalar() or 0)

    last_note = db.query(func.max(models.CheckIn.ts)).filter(*note_f).scalar()
    last_stat = db.query(func.max(models.StatusChange.ts)).filter(*stat_f).scalar()
    latest = max([t for t in (last_note, last_stat) if t], default=None)
    return {"count": count,
            "latest": latest.isoformat(timespec="seconds") if latest else None}


# --------------------------------------------------------------- excel import
IMPORT_FIELDS = ["name", "asset_class", "department", "owner",
                 "vendor", "location", "cost", "serial_number",
                 "date_purchased", "status", "notes"]


def _clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _coerce_class(v):
    s = _clean_str(v)
    return (s.lstrip("*").strip() or None) if s else None


def _coerce_cost(v):
    if v is None or v == "":
        return None, None
    if isinstance(v, (int, float)):
        return float(v), None
    s = str(v).strip().replace("$", "").replace(",", "").strip()
    if not s:
        return None, None
    try:
        return float(s), None
    except ValueError:
        return None, f"cost '{v}' isn't a number — left blank"


def _coerce_date(v):
    if v is None or v == "":
        return None, None
    if isinstance(v, _dt):               # openpyxl returns datetime for real date cells
        return v.date(), None
    if isinstance(v, date_type):
        return v, None
    s = str(v).strip()
    try:
        return _parse_date(s), None      # yyyy-mm-dd fast path
    except Exception:
        for fmt in ("%d-%b-%y", "%d-%b-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%y"):
            try:
                return _dt.strptime(s, fmt).date(), None
            except ValueError:
                continue
        return None, f"date '{v}' unreadable — left blank"


def _coerce_status(v):
    s = _clean_str(v)
    if s is None:
        return models.DEFAULT_STATUS, None
    for st in models.STATUSES:
        if st.lower() == s.lower():
            return st, None
    return models.DEFAULT_STATUS, f"status '{v}' not recognized — set to {models.DEFAULT_STATUS}"


def _load_ws(raw, sheet, header_row):
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    names = wb.sheetnames
    ws = wb[sheet] if sheet in names else wb[names[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return names, ws.title, rows


def _headers_and_data(rows, header_row):
    if header_row < 1 or header_row > len(rows):
        raise HTTPException(400, f"Header row {header_row} is out of range (the sheet has {len(rows)} rows).")
    headers = [(str(c).strip() if c is not None else "") for c in rows[header_row - 1]]
    return headers, rows[header_row:]


def _parse_import(raw, sheet, header_row, mapping, dept_lookup):
    _, _, rows = _load_ws(raw, sheet, header_row)
    headers, data = _headers_and_data(rows, header_row)
    col_index = {h: i for i, h in enumerate(headers) if h}

    # every mapped column must actually exist in the chosen header row
    for field, col in mapping.items():
        if col and col not in col_index:
            raise HTTPException(400, f"Column '{col}' (mapped to {field}) isn't in the header row.")

    assets, warnings, skipped = [], [], 0
    for i, row in enumerate(data):
        def cell(field):
            col = mapping.get(field)
            if not col:
                return None
            idx = col_index.get(col)
            return row[idx] if (idx is not None and idx < len(row)) else None

        name = _clean_str(cell("name"))
        if not name:                      # blank name => banner/spacer row => skip
            skipped += 1
            continue

        f, msgs = {"name": name}, []
        for plain in ("owner", "vendor", "location", "serial_number", "notes"):
            f[plain] = _clean_str(cell(plain))
        f["asset_class"] = _coerce_class(cell("asset_class"))
        f["cost"], w = _coerce_cost(cell("cost"));               msgs += [w] if w else []
        f["date_purchased"], w = _coerce_date(cell("date_purchased")); msgs += [w] if w else []
        if mapping.get("status"):
            f["status"], w = _coerce_status(cell("status"));     msgs += [w] if w else []
        else:
            f["status"] = models.DEFAULT_STATUS

        dept_name = _clean_str(cell("department"))
        if dept_name:
            did = dept_lookup.get(dept_name.lower())
            if did:
                f["department_id"] = did
            else:
                msgs.append(f"department '{dept_name}' not found — left blank")

        if msgs:
            warnings.append({"row": header_row + 1 + i, "messages": msgs})
        assets.append(f)

    return assets, warnings, skipped, len(data)


def _make_asset(db, f, user_email):
    status = f.get("status") or models.DEFAULT_STATUS
    a = models.Asset(
        name=f["name"], location=f.get("location"),
        department_id=f.get("department_id") or None,
        owner=f.get("owner"), vendor=f.get("vendor"),
        serial_number=f.get("serial_number"), asset_class=f.get("asset_class"),
        cost=f.get("cost"),
        date_purchased=f.get("date_purchased"), notes=f.get("notes"),
        status=status, status_note=None, status_updated_at=models.utcnow())
    db.add(a); db.flush()
    db.add(models.StatusChange(asset_id=a.id, old_status=None, new_status=status,
                               note=None, changed_by=user_email))
    return a


def _preview_row(f):
    d = dict(f)
    if d.get("date_purchased"):
        d["date_purchased"] = d["date_purchased"].isoformat()
    return d


@router.post("/import/inspect")
async def import_inspect(file: UploadFile = File(...), header_row: int = Form(1),
                         sheet: str = Form(""), user=Depends(require_owner)):
    raw = await file.read()
    try:
        names, title, rows = _load_ws(raw, sheet, header_row)
    except Exception as e:
        raise HTTPException(400, f"Couldn't read this file as .xlsx ({e}). Only .xlsx is supported.")
    headers, data = _headers_and_data(rows, header_row)
    sample = [[("" if c is None else str(c)) for c in row][:len(headers)] for row in data[:5]]
    return {"sheets": names, "sheet": title, "header_row": header_row,
            "columns": [h for h in headers if h], "sample": sample,
            "data_rows": len(data), "fields": IMPORT_FIELDS}


@router.post("/import/preview")
async def import_preview(file: UploadFile = File(...), header_row: int = Form(1),
                         sheet: str = Form(""), mapping: str = Form(...),
                         db: Session = Depends(get_db), user=Depends(require_owner)):
    m = json.loads(mapping)
    if not m.get("name"):
        raise HTTPException(400, "Map a column to Name — it's the only required field.")
    raw = await file.read()
    dept_lookup = {d.name.lower(): d.id for d in db.query(models.Department).all()}
    assets, warnings, skipped, total = _parse_import(raw, sheet, header_row, m, dept_lookup)
    return {"total_rows": total, "will_import": len(assets), "skipped_blank": skipped,
            "warning_count": len(warnings), "warnings": warnings[:300],
            "preview": [_preview_row(a) for a in assets[:10]]}


@router.post("/import/commit")
async def import_commit(file: UploadFile = File(...), header_row: int = Form(1),
                        sheet: str = Form(""), mapping: str = Form(...),
                        db: Session = Depends(get_db), user=Depends(require_owner)):
    m = json.loads(mapping)
    if not m.get("name"):
        raise HTTPException(400, "Map a column to Name — it's the only required field.")
    raw = await file.read()
    dept_lookup = {d.name.lower(): d.id for d in db.query(models.Department).all()}
    assets, warnings, skipped, total = _parse_import(raw, sheet, header_row, m, dept_lookup)
    try:
        for f in assets:
            _make_asset(db, f, user.email)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(400, f"Import failed — nothing was saved ({e}).")
    return {"created": len(assets), "skipped_blank": skipped, "warning_count": len(warnings)}


# --------------------------------------------------------------- analytics
@router.get("/analytics")
def analytics(exclude_missing: bool = False,
              db: Session = Depends(get_db), user=Depends(current_user)):
    """Numbers for the dashboard. Everything is computed live from the same
    tables the rest of the app uses - there is no separate reporting store.

    exclude_missing drops the "(none)" buckets from the breakdown charts so
    incomplete legacy data doesn't dominate them."""
    assets = db.query(models.Asset).filter((models.Asset.active == True)).all()
    total = len(assets)
    now = models.utcnow()

    # Data-quality counts, always reported so the gaps are visible even when
    # their "(none)" bars are hidden.
    missing = {
        "location": sum(1 for a in assets if not (a.location or "").strip()),
        "department": sum(1 for a in assets if not a.department_id),
        "cost": sum(1 for a in assets if a.cost is None),
        "date_purchased": sum(1 for a in assets if not a.date_purchased),
    }

    by_status, by_dept, by_location, by_class, by_equipment = {}, {}, {}, {}, {}
    value_total, value_by_dept, costed = 0.0, {}, 0
    for a in assets:
        by_status[a.status] = by_status.get(a.status, 0) + 1
        dname = a.department.name if a.department else "(none)"
        by_dept[dname] = by_dept.get(dname, 0) + 1
        by_location[a.location or "(none)"] = by_location.get(a.location or "(none)", 0) + 1
        by_class[a.asset_class or "(none)"] = by_class.get(a.asset_class or "(none)", 0) + 1
        if a.cost is not None:
            value_total += float(a.cost)
            costed += 1
            value_by_dept[dname] = value_by_dept.get(dname, 0.0) + float(a.cost)

    qualified = by_status.get("Qualified", 0)

    # Average days from an asset first appearing to first being marked Qualified.
    # Uses the status log, so it reflects what actually happened rather than
    # anything a person typed in.
    first_seen, first_qualified = {}, {}
    for sc in db.query(models.StatusChange).order_by(models.StatusChange.ts).all():
        if sc.asset_id not in first_seen:
            first_seen[sc.asset_id] = sc.ts
        if sc.new_status == "Qualified" and sc.asset_id not in first_qualified:
            first_qualified[sc.asset_id] = sc.ts
    spans = [(first_qualified[i] - first_seen[i]).total_seconds() / 86400.0
             for i in first_qualified if i in first_seen
             and first_qualified[i] >= first_seen[i]]
    spans.sort()
    avg_days = round(sum(spans) / len(spans), 1) if spans else None
    median_days = round(spans[len(spans) // 2], 1) if spans else None

    # Assets still waiting to be qualified, and how long they've been waiting.
    pending = [a for a in assets if a.status == "Purchased"]
    waiting = []
    for a in pending:
        started = first_seen.get(a.id) or a.created_at
        if started:
            waiting.append((now - started).total_seconds() / 86400.0)
    avg_waiting = round(sum(waiting) / len(waiting), 1) if waiting else None

    # Check-in activity for the last 30 days.
    since = now - datetime.timedelta(days=30)
    daily = {}
    for c in db.query(models.CheckIn).filter(models.CheckIn.ts >= since).all():
        key = c.ts.date().isoformat()
        daily[key] = daily.get(key, 0) + 1
    checkins_30d = [{"date": (since + datetime.timedelta(days=i)).date().isoformat(),
                     "count": daily.get((since + datetime.timedelta(days=i)).date().isoformat(), 0)}
                    for i in range(31)]

    # Coverage: never checked in, and stale (>90 days).
    last_seen = {}
    for c in db.query(models.CheckIn).all():
        if c.asset_id not in last_seen or c.ts > last_seen[c.asset_id]:
            last_seen[c.asset_id] = c.ts
    never = sum(1 for a in assets if a.id not in last_seen)
    stale = sum(1 for a in assets if a.id in last_seen
                and (now - last_seen[a.id]).days > 90)

    def top(d, n=8):
        items = d.items()
        if exclude_missing:
            items = [(k, v) for k, v in items if k != "(none)"]
        return [{"label": k, "count": v} for k, v in
                sorted(items, key=lambda kv: -kv[1])[:n]]

    def top_val(d, n=8):
        items = [(k, v) for k, v in d.items() if not (exclude_missing and k == "(none)")]
        return [{"label": k, "value": round(v, 2)} for k, v in
                sorted(items, key=lambda kv: -kv[1])[:n]]

    return {
        "total": total,
        "qualified": qualified,
        "qualified_pct": round(qualified / total * 100, 1) if total else 0,
        "pending_qualification": len(pending),
        "avg_days_to_qualify": avg_days,
        "median_days_to_qualify": median_days,
        "qualified_sample": len(spans),
        "avg_days_waiting": avg_waiting,
        "value_total": round(value_total, 2),
        "value_missing_cost": total - costed,
        "never_checked_in": never,
        "stale_90d": stale,
        "missing": missing,
        "excluded_missing": exclude_missing,
        "by_status": [{"label": s, "count": by_status.get(s, 0)} for s in models.STATUSES
                      if by_status.get(s)],
        "by_department": top(by_dept),
        "by_location": top(by_location),
        "by_class": top(by_class),
        "value_by_department": top_val(value_by_dept),
        "checkins_30d": checkins_30d,
    }


# --------------------------------------------------------------- assets
def _brief(a: models.Asset):
    last = a.checkins[0] if a.checkins else None
    last_status = a.status_changes[0] if a.status_changes else None
    return {
        "id": a.id, "code": a.code, "asset_code": a.code, "name": a.name,
        "location": a.location,
        "department_id": a.department_id,
        "department": a.department.name if a.department else None,
        "owner": a.owner, "vendor": a.vendor, "serial_number": a.serial_number,
        "asset_class": a.asset_class,
        "cost": a.cost,
        "date_purchased": a.date_purchased.isoformat() if a.date_purchased else None,
        "notes": a.notes, "status": a.status, "status_note": a.status_note,
        "status_updated_at": a.status_updated_at.isoformat(timespec="seconds")
                              if a.status_updated_at else None,
        "status_updated_by": last_status.changed_by if last_status else None,
        "previous_status": last_status.old_status if last_status else None,
        "status_change_count": len(a.status_changes),
        "active": a.active,
        "checked_out": bool(a.checked_out),
        "holder_department_id": a.holder_department_id,
        "holder_department": a.holder_department.name if a.holder_department else None,
        "holder_person": a.holder_person,
        "checked_out_at": a.checked_out_at.isoformat(timespec="seconds") if a.checked_out_at else None,
        "due_back": a.due_back.isoformat() if a.due_back else None,
        "reminders_due": _reminders_due_count(a),
        "created_at": a.created_at.isoformat(timespec="seconds"),
        "checkin_count": len(a.checkins),
        "last_checkin": last.ts.isoformat(timespec="seconds") if last else None,
        "last_checked_in_by": last.checked_in_by if last else None,
    }


def _apply_status_change(db: Session, a: models.Asset, new_status: str,
                         note: Optional[str], changed_by: Optional[str]):
    """Updates the asset's current status and logs the change, only if it
    actually changed."""
    if new_status == a.status and (note or None) == (a.status_note or None):
        return
    became_qualified = new_status == "Qualified" and a.status != "Qualified"
    db.add(models.StatusChange(asset_id=a.id, old_status=a.status,
                               new_status=new_status, note=note, changed_by=changed_by))
    a.status = new_status
    a.status_note = note
    a.status_updated_at = models.utcnow()
    if became_qualified:
        notify_qualified(db, a, changed_by, note)


def _filter_assets(assets, q: Optional[str]):
    if not q:
        return assets
    ql = q.lower()
    def hit(a):
        fields = [a.name, a.location, a.serial_number, a.code,
                  a.owner, a.vendor, a.status,
                  a.department.name if a.department else ""]
        return any(ql in (f or "").lower() for f in fields)
    return [a for a in assets if hit(a)]


def _dup_key(a: models.Asset):
    """Assets are 'the same' when the identifying fields all match. Serial
    number alone counts too (a serial is meant to be unique). Used only to
    flag possible duplicates for the user — nothing is blocked or merged."""
    name = (a.name or "").strip().lower()
    serial = (a.serial_number or "").strip().lower()
    keys = []
    if serial:
        keys.append(("serial", serial))
    keys.append(("full", name, (a.vendor or "").strip().lower(),
                 (a.department_id or 0), round(a.cost, 2) if a.cost is not None else None,
                 a.date_purchased.isoformat() if a.date_purchased else None))
    return keys


def _duplicate_ids(assets):
    """Returns the set of asset ids that share a duplicate key with at least
    one other asset."""
    from collections import defaultdict
    buckets = defaultdict(list)
    for a in assets:
        for k in _dup_key(a):
            buckets[k].append(a.id)
    dupes = set()
    for k, ids in buckets.items():
        if len(ids) > 1:
            dupes.update(ids)
    return dupes


@router.get("/assets")
def list_assets(q: Optional[str] = None, limit: int = 0, offset: int = 0,
                active_only: bool = True,
                db: Session = Depends(get_db), user=Depends(current_user)):
    query = db.query(models.Asset)
    if active_only:
        query = query.filter((models.Asset.active == True))
    all_matching = _filter_assets(query.order_by(models.Asset.name).all(), q)

    # Duplicate detection runs across the whole (filtered) set, not just the
    # current page, so a dupe is flagged even if its twin is on another page.
    dupes = _duplicate_ids(all_matching)

    total = len(all_matching)
    page = all_matching[offset: offset + limit] if limit else all_matching
    rows = []
    for a in page:
        b = _brief(a)
        b["is_duplicate"] = a.id in dupes
        rows.append(b)
    # Backwards compatible: with no limit, still return a bare list.
    if not limit:
        return rows
    return {"total": total, "offset": offset, "limit": limit,
            "duplicate_count": len(dupes), "assets": rows}


# ------------------------------------------------------------- custody
@router.post("/assets/{asset_id}/checkout")
def checkout_asset(asset_id: int, body: dict, db: Session = Depends(get_db),
                   user=Depends(require_support)):
    """Record that an asset has physically left its home, borrowed by a
    department (required) with an optional person. Does not touch the owning
    department or the lifecycle status - purely a note of who has it now."""
    a = db.get(models.Asset, asset_id)
    if not a or not a.active:
        raise HTTPException(404, "Asset not found")
    dept_id = body.get("department_id")
    if not dept_id:
        raise HTTPException(400, "A borrowing department is required to check out.")
    dept = db.get(models.Department, int(dept_id))
    if not dept:
        raise HTTPException(400, "Unknown department.")
    a.checked_out = True
    a.holder_department_id = dept.id
    a.holder_person = (body.get("person") or "").strip() or None
    a.checked_out_at = models.utcnow()
    a.due_back = _parse_date(body.get("due_back"))
    db.add(models.CustodyEvent(asset_id=a.id, action="out", department_id=dept.id,
                               person=a.holder_person, due_back=a.due_back,
                               note=(body.get("note") or "").strip() or None,
                               by_user=user.email))
    db.commit(); db.refresh(a)
    return _brief(a)


@router.post("/assets/{asset_id}/checkin_custody")
def checkin_custody(asset_id: int, body: dict, db: Session = Depends(get_db),
                    user=Depends(require_support)):
    """Record that a checked-out asset has physically come back."""
    a = db.get(models.Asset, asset_id)
    if not a or not a.active:
        raise HTTPException(404, "Asset not found")
    if not a.checked_out:
        raise HTTPException(400, "This asset isn't checked out.")
    db.add(models.CustodyEvent(asset_id=a.id, action="in",
                               department_id=a.holder_department_id, person=a.holder_person,
                               note=(body.get("note") or "").strip() or None,
                               by_user=user.email))
    a.checked_out = False
    a.holder_department_id = None
    a.holder_person = None
    a.checked_out_at = None
    a.due_back = None
    db.commit(); db.refresh(a)
    return _brief(a)


@router.get("/assets/{asset_id}/custody")
def custody_history(asset_id: int, db: Session = Depends(get_db), user=Depends(current_user)):
    rows = (db.query(models.CustodyEvent)
            .filter(models.CustodyEvent.asset_id == asset_id)
            .order_by(models.CustodyEvent.ts.desc()).all())
    out = []
    for e in rows:
        dept = db.get(models.Department, e.department_id) if e.department_id else None
        out.append({"ts": e.ts.isoformat(timespec="seconds"), "action": e.action,
                    "department": dept.name if dept else None, "person": e.person,
                    "due_back": e.due_back.isoformat() if e.due_back else None,
                    "note": e.note, "by_user": e.by_user})
    return {"events": out}


# ------------------------------------------------------------- reminders
def _reminders_due_count(a: models.Asset) -> int:
    """How many active reminders on this asset are due today or overdue."""
    from datetime import date as _d
    today = _d.today()
    return sum(1 for r in a.reminders_list
               if r.active and r.next_due and r.next_due <= today)


def _reminder_brief(r: models.Reminder):
    from datetime import date as _d
    today = _d.today()
    days = (r.next_due - today).days if r.next_due else None
    state = "none"
    if r.active and r.next_due:
        state = "overdue" if days < 0 else ("due" if days <= 0 else
                ("soon" if days <= 14 else "upcoming"))
    return {"id": r.id, "asset_id": r.asset_id, "label": r.label, "kind": r.kind,
            "interval_days": r.interval_days,
            "next_due": r.next_due.isoformat() if r.next_due else None,
            "days_until": days, "state": state, "active": r.active,
            "note": r.note, "last_done_at": r.last_done_at.isoformat() if r.last_done_at else None}


@router.get("/assets/{asset_id}/reminders")
def list_reminders(asset_id: int, db: Session = Depends(get_db), user=Depends(current_user)):
    rows = (db.query(models.Reminder)
            .filter(models.Reminder.asset_id == asset_id)
            .order_by(models.Reminder.active.desc(), models.Reminder.next_due).all())
    return {"reminders": [_reminder_brief(r) for r in rows]}


def _create_one_reminder(db, asset_id, body, user_email):
    label = (body.get("label") or "").strip()
    if not label:
        raise HTTPException(400, "A reminder needs a label.")
    kind = body.get("kind") if body.get("kind") in ("interval", "date") else "interval"
    r = models.Reminder(asset_id=asset_id, label=label, kind=kind,
                        note=(body.get("note") or "").strip() or None,
                        created_by=user_email)
    if kind == "interval":
        try:
            n = int(body.get("interval_days"))
        except (TypeError, ValueError):
            raise HTTPException(400, "Recurring reminders need a number of days.")
        if n < 1:
            raise HTTPException(400, "Interval must be at least 1 day.")
        r.interval_days = n
        start = _parse_date(body.get("next_due")) if body.get("next_due") else None
        r.next_due = start or _first_due(n)
    else:
        d = _parse_date(body.get("next_due"))
        if not d:
            raise HTTPException(400, "A one-off reminder needs a date.")
        r.next_due = d
    db.add(r)
    return r


def _first_due(interval_days):
    from datetime import date as _d, timedelta
    return _d.today() + timedelta(days=interval_days)


@router.post("/assets/{asset_id}/reminders")
def add_reminder(asset_id: int, body: dict, db: Session = Depends(get_db),
                 user=Depends(require_support)):
    a = db.get(models.Asset, asset_id)
    if not a:
        raise HTTPException(404, "Asset not found")
    r = _create_one_reminder(db, asset_id, body, user.email)
    db.commit(); db.refresh(r)
    return _reminder_brief(r)


@router.post("/reminders/bulk")
def add_reminder_bulk(body: dict, db: Session = Depends(get_db), user=Depends(require_support)):
    """Attach the same reminder to many assets at once (select-all support)."""
    ids = [int(i) for i in (body.get("asset_ids") or [])]
    if not ids:
        raise HTTPException(400, "No assets selected.")
    made = 0
    for aid in ids:
        if db.get(models.Asset, aid):
            _create_one_reminder(db, aid, body, user.email)
            made += 1
    db.commit()
    return {"created": made}


@router.post("/reminders/{reminder_id}/done")
def complete_reminder(reminder_id: int, body: dict, db: Session = Depends(get_db),
                      user=Depends(require_support)):
    """Mark a reminder done. Recurring ones roll forward from the due date;
    one-off ones are switched off."""
    r = db.get(models.Reminder, reminder_id)
    if not r:
        raise HTTPException(404, "Reminder not found")
    from datetime import date as _d, timedelta
    done_on = _parse_date(body.get("done_on")) or _d.today()
    r.last_done_at = done_on
    db.add(models.ReminderLog(reminder_id=r.id, done_on=done_on, by_user=user.email,
                              note=(body.get("note") or "").strip() or None))
    if r.kind == "interval" and r.interval_days:
        base = r.next_due or done_on
        # Roll forward from the due date, catching up past any missed cycles so
        # the next due date is always in the future on a clean grid.
        while base <= _d.today():
            base = base + timedelta(days=r.interval_days)
        r.next_due = base
    else:
        r.active = False
    db.commit(); db.refresh(r)
    return _reminder_brief(r)


@router.delete("/reminders/{reminder_id}")
def delete_reminder(reminder_id: int, db: Session = Depends(get_db), user=Depends(require_support)):
    r = db.get(models.Reminder, reminder_id)
    if not r:
        raise HTTPException(404, "Reminder not found")
    db.delete(r); db.commit()
    return {"deleted": True}


@router.get("/reminders/upcoming")
def reminders_upcoming(db: Session = Depends(get_db), user=Depends(current_user)):
    """The notification center: every active reminder, grouped by urgency."""
    rows = (db.query(models.Reminder)
            .filter((models.Reminder.active == True), models.Reminder.next_due.isnot(None))
            .order_by(models.Reminder.next_due).all())
    groups = {"overdue": [], "due": [], "soon": [], "upcoming": []}
    for r in rows:
        a = db.get(models.Asset, r.asset_id)
        if not a or not a.active:
            continue
        item = _reminder_brief(r)
        item["asset_name"] = a.name
        item["asset_code"] = a.code
        bucket = item["state"] if item["state"] in groups else "upcoming"
        if bucket == "none":
            continue
        groups[bucket].append(item)
    due_now = len(groups["overdue"]) + len(groups["due"])
    return {"groups": groups,
            "counts": {k: len(v) for k, v in groups.items()},
            "due_now": due_now}


@router.post("/assets/bulk-status")
def bulk_status(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    """Set the same status on many assets at once (e.g. Purchased -> Qualified
    for a whole batch). Each change is logged individually and fires the same
    side effects (Qualified email) as a single change."""
    status = (body.get("status") or "").strip()
    ids = [int(i) for i in (body.get("ids") or [])]
    if status not in models.STATUSES:
        raise HTTPException(400, "Unknown status.")
    if not ids:
        raise HTTPException(400, "No assets selected.")
    rows = db.query(models.Asset).filter(models.Asset.id.in_(ids)).all()
    for a in rows:
        _apply_status_change(db, a, status, None, user.email)
    db.commit()
    return {"updated": len(rows), "status": status}


@router.post("/assets/bulk")
def bulk_assets(body: dict, db: Session = Depends(get_db), user=Depends(require_support)):
    """Apply one action to many assets at once, for the select-all checkboxes.

    action: "deactivate" (reversible, keeps history) | "reactivate"
            | "delete" (permanent, also removes that asset's history)
    """
    action = (body.get("action") or "").strip()
    ids = [int(i) for i in (body.get("ids") or [])]
    if action not in ("deactivate", "reactivate", "delete"):
        raise HTTPException(400, "action must be deactivate, reactivate or delete")
    if not ids:
        raise HTTPException(400, "No assets selected.")
    rows = db.query(models.Asset).filter(models.Asset.id.in_(ids)).all()
    for a in rows:
        if action == "delete":
            db.delete(a)          # cascade removes its check-ins + status log
        else:
            a.active = (action == "reactivate")
    db.commit()
    return {"action": action, "affected": len(rows)}


@router.get("/assets/duplicates")
def list_duplicates(db: Session = Depends(get_db), user=Depends(current_user)):
    """Groups of assets that look like duplicates of each other."""
    from collections import defaultdict
    assets = db.query(models.Asset).filter((models.Asset.active == True)).all()
    buckets = defaultdict(list)
    for a in assets:
        for k in _dup_key(a):
            buckets[k].append(a)
    groups = []
    seen = set()
    for k, members in buckets.items():
        if len(members) < 2:
            continue
        ids = tuple(sorted(m.id for m in members))
        if ids in seen:
            continue
        seen.add(ids)
        reason = "same serial number" if k[0] == "serial" else "same name, vendor, dept, cost & date"
        groups.append({"reason": reason,
                       "assets": [{"id": m.id, "name": m.name,
                                   "asset_code": m.code,
                                   "serial_number": m.serial_number,
                                   "department": m.department.name if m.department else None,
                                   "created_at": m.created_at.isoformat(timespec="seconds")}
                                  for m in sorted(members, key=lambda x: x.id)]})
    return {"groups": groups, "count": len(groups)}


@router.get("/assets/export.xlsx")
def export_assets_xlsx(q: Optional[str] = None, db: Session = Depends(get_db),
                       user=Depends(current_user)):
    # Must stay registered before /assets/{asset_id} - otherwise FastAPI
    # tries to parse "export.xlsx" as an integer asset_id and 422s.
    assets = _filter_assets(
        db.query(models.Asset).order_by(models.Asset.name).all(), q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    headers = ["Name", "Asset Code", "Department", "Owner", "Vendor",
              "Location", "Serial Number", "Cost", "Date Purchased",
              "Status", "Status Note", "Last Check-In", "Check-In Count", "Notes"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="0B1F3A")
    for a in assets:
        last = a.checkins[0] if a.checkins else None
        ws.append([
            a.name, a.code, a.department.name if a.department else "",
            a.owner or "", a.vendor or "", a.location or "", a.serial_number or "",
            a.cost, a.date_purchased.isoformat() if a.date_purchased else "",
            a.status, a.status_note or "",
            last.ts.strftime("%Y-%m-%d %H:%M") if last else "",
            len(a.checkins), a.notes or "",
        ])
    for i, w in enumerate([26, 14, 16, 16, 16, 18, 16, 10, 14, 14, 22, 20, 14, 30], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="assets.xlsx"'})


@router.post("/assets")
def create_asset(body: dict, db: Session = Depends(get_db), user=Depends(require_support)):
    if not body.get("name"):
        raise HTTPException(400, "Name is required")
    status = body.get("status") or models.DEFAULT_STATUS
    if status not in models.STATUSES:
        raise HTTPException(400, f"Status must be one of: {', '.join(models.STATUSES)}")
    a = models.Asset(name=body["name"],
                     location=body.get("location"),
                     department_id=body.get("department_id") or None,
                     owner=body.get("owner"), vendor=body.get("vendor"),
                     serial_number=body.get("serial_number"),
                     asset_class=body.get("asset_class"),
                     cost=_parse_cost(body.get("cost")),
                     date_purchased=_parse_date(body.get("date_purchased")),
                     notes=body.get("notes"),
                     status=status, status_note=body.get("status_note"),
                     status_updated_at=models.utcnow())
    db.add(a); db.flush()
    db.add(models.StatusChange(asset_id=a.id, old_status=None, new_status=status,
                               note=body.get("status_note"), changed_by=user.email))
    db.commit(); db.refresh(a)
    return _brief(a)


@router.get("/assets/{asset_id}")
def get_asset(asset_id: int, db: Session = Depends(get_db), user=Depends(current_user)):
    a = db.get(models.Asset, asset_id)
    if not a:
        raise HTTPException(404, "Asset not found")
    d = _brief(a)
    d["checkin_url"] = checkin_url(base_url(db), a.code)
    history = []
    for c in a.checkins[:200]:
        history.append({"type": "scan", "ts": c.ts.isoformat(),
                        "id": c.id, "checked_in_by": c.checked_in_by, "note": c.note,
                        "lat": c.lat, "lon": c.lon, "accuracy_m": c.accuracy_m,
                        "source": c.source})
    status_history = []
    for s in a.status_changes[:200]:
        entry = {"type": "status_change", "ts": s.ts.isoformat(),
                 "id": s.id, "old_status": s.old_status, "new_status": s.new_status,
                 "note": s.note, "changed_by": s.changed_by}
        history.append(entry)
        status_history.append(entry.copy())
    history.sort(key=lambda h: h["ts"], reverse=True)
    status_history.sort(key=lambda h: h["ts"], reverse=True)
    d["history"] = history[:250]
    d["status_history"] = status_history
    return d


@router.put("/assets/{asset_id}")
def update_asset(asset_id: int, body: dict, db: Session = Depends(get_db), user=Depends(require_support)):
    a = db.get(models.Asset, asset_id)
    if not a:
        raise HTTPException(404, "Asset not found")
    if "status" in body and body["status"] not in models.STATUSES and body["status"] != a.status:
        raise HTTPException(400, f"Status must be one of: {', '.join(models.STATUSES)}")
    if "status" in body:
        _apply_status_change(db, a, body["status"], body.get("status_note"), user.email)
    for f in ("name", "location", "owner", "vendor",
             "serial_number", "asset_class", "notes", "active"):
        if f in body:
            setattr(a, f, body[f])
    if "cost" in body:
        a.cost = _parse_cost(body["cost"])
    if "date_purchased" in body:
        a.date_purchased = _parse_date(body.get("date_purchased"))
    if "department_id" in body:
        a.department_id = body["department_id"] or None
    # If the asset never got a code (missing inputs at creation) and the edit
    # has now filled the gaps, assign one. Existing codes are left untouched.
    db.commit(); db.refresh(a)
    return _brief(a)


@router.delete("/assets/{asset_id}")
def delete_asset(asset_id: int, db: Session = Depends(get_db), user=Depends(require_owner)):
    a = db.get(models.Asset, asset_id)
    if not a:
        raise HTTPException(404, "Asset not found")
    db.delete(a); db.commit()
    return {"deleted": True}


@router.post("/assets/{asset_id}/manual-checkin")
def manual_checkin(asset_id: int, db: Session = Depends(get_db), user=Depends(current_user)):
    """
    A deliberate, confirmed check-in logged from inside the dashboard (not
    the public scan page). Requires the person to already be signed in, and
    the frontend confirms with the user before calling this - so this is
    never fired by simply opening or viewing an asset.
    """
    a = db.get(models.Asset, asset_id)
    if not a or not a.active:
        raise HTTPException(404, "Asset not found")
    c = models.CheckIn(asset_id=a.id, checked_in_by=user.full_name,
                      note=None, source="manual_admin")
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "ts": c.ts.isoformat(timespec="seconds")}


# --------------------------------------------------------------- QR / barcode / label
@router.get("/assets/{asset_id}/qr.png")
def qr_png(asset_id: int, db: Session = Depends(get_db)):
    # No login required (see module docstring): <img src> sends no auth header.
    a = db.get(models.Asset, asset_id)
    if not a:
        raise HTTPException(404, "Asset not found")
    png = make_qr_png(checkin_url(base_url(db), a.code))
    return StreamingResponse(io.BytesIO(png), media_type="image/png")


@router.get("/assets/{asset_id}/barcode.png")
def barcode_png(asset_id: int, text: int = 1, db: Session = Depends(get_db)):
    # No login required, same reasoning as qr_png above.
    # text=0 drops the human-readable string under the bars (used by the label,
    # which prints the asset code as text itself).
    a = db.get(models.Asset, asset_id)
    if not a:
        raise HTTPException(404, "Asset not found")
    png = make_barcode_png(a.code, write_text=bool(text))
    return StreamingResponse(io.BytesIO(png), media_type="image/png")


@router.get("/assets/{asset_id}/label", response_class=HTMLResponse)
def label(asset_id: int, db: Session = Depends(get_db)):
    """Printable asset tag, sized for Brady M5-29-428 stock (1.44in x 0.5in).

    Layout: Frore mark top-left, PROPERTY OF / company / asset code centred,
    Code128 barcode beneath them, QR on the right. The barcode is requested
    with text=0 because the code already appears above it as text.
    """
    a = db.get(models.Asset, asset_id)
    if not a:
        raise HTTPException(404, "Asset not found")
    display_code = a.code
    company = company_name(db)

    # Shrink the type a step when the code or company name is long, so the
    # label degrades gracefully instead of overflowing the stock.
    n = len(display_code)
    code_pt = 5.2 if n <= 18 else (4.6 if n <= 24 else 4.0)
    c = len(company)
    brand_pt = 7.2 if c <= 15 else (6.1 if c <= 21 else 5.0)

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Label - {_esc(a.name)}</title>
<style>
  @page {{ size: 1.44in 0.5in; margin: 0; }}
  body {{ font-family: Arial, Helvetica, sans-serif; background: #eef2f7;
          margin: 0; padding: 40px; text-align: center; }}
  .tag {{ position: relative; box-sizing: border-box; display: inline-block;
          width: 1.44in; height: 0.5in; background: #fff; overflow: hidden; }}
  .logo {{ position: absolute; top: 0.012in; left: 0.012in;
           width: 0.075in; height: 0.075in; object-fit: contain; }}
  .qr {{ position: absolute; top: 0.045in; right: 0.022in;
         width: 0.42in; height: 0.42in; display: block; }}
  .mid {{ position: absolute; left: 0.022in; right: 0.455in; top: 0.028in; text-align: center; }}
  .prop {{ font-size: 4pt; font-weight: 400; line-height: 1;
           letter-spacing: .55pt; text-transform: uppercase; color: #000; }}
  .brand {{ font-size: {brand_pt}pt; font-weight: 600; line-height: 1.15;
            letter-spacing: -.05pt; margin-top: .022in; color: #000; white-space: nowrap; }}
  .code {{ font-size: {code_pt}pt; font-weight: 500; line-height: 1;
           letter-spacing: .35pt; margin-top: .026in; color: #000; white-space: nowrap; }}
  .bar {{ display: block; width: .76in; height: .1in; margin: .022in auto 0; }}
  .hint {{ color: #7284a0; font-size: 12px; margin-top: 20px; line-height: 1.5; }}
  .hint b {{ color: #45536b; }}
  .note {{ color: #b4622a; font-size: 12px; font-weight: 700; margin-bottom: 14px; }}
  .btn {{ font: 500 13px Arial; padding: 7px 13px; margin-top: 16px; cursor: pointer;
          background: #fff; color: #16233a; border: 1px solid #c3cfe0; border-radius: 6px; }}
  .btn:hover {{ border-color: #2f7de0; color: #2f7de0; }}
  @media print {{
    body {{ background: #fff; padding: 0; margin: 0; }}
    .hint, .note, .btn {{ display: none; }}
    .tag {{ zoom: 1 !important; }}
  }}
</style></head><body>
  <div class="note" id="szNote">*not actual size</div>
  <div class="tag" id="tag" style="zoom:5">
    <img class="logo" src="/static/frore-logo.png" alt="" />
    <div class="mid">
      <div class="prop">PROPERTY OF</div>
      <div class="brand">{_esc(company)}</div>
      <div class="code">{_esc(display_code)}</div>
      <img class="bar" src="/api/assets/{a.id}/barcode.png?text=0" alt="" />
    </div>
    <img class="qr" src="/api/assets/{a.id}/qr.png" alt="QR code" />
  </div>
  <div><button class="btn" id="szBtn" onclick="toggleSize()">Show at actual size</button></div>
  <div class="hint">Real label is <b>1.44in &times; 0.5in</b> (Brady M5-29-428).<br />
    Press <b>Ctrl+P / Cmd+P</b> to print &mdash; it always prints at true size, whatever the preview shows.<br />
    Set printer scale to <b>100%</b>, not "fit to page".</div>
<script>
  var big = true;
  function toggleSize() {{
    big = !big;
    document.getElementById("tag").style.zoom = big ? 5 : 1;
    document.getElementById("szNote").textContent = big ? "*not actual size" : "actual size";
    document.getElementById("szBtn").textContent = big ? "Show at actual size" : "Show enlarged";
  }}
</script>
</body></html>"""


# --------------------------------------------------------------- inventory scan
@router.post("/scan-action")
def scan_action(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    """One scan, one action - the engine behind the multi-mode scan page.

    mode="lookup"   : just find the asset and return it (no change)
    mode="checkin"  : log an inventory check-in (any signed-in user)
    mode="checkout" : check it out to department_id (needs lab_support)
    mode="return"   : check it back in from custody (needs lab_support)
    """
    raw = (body.get("code") or "").strip().upper()
    mode = body.get("mode") or "checkin"
    if not raw:
        raise HTTPException(400, "No code provided")
    a = db.query(models.Asset).filter(models.Asset.code == raw).first()
    if not a or not a.active:
        raise HTTPException(404, f"No active asset matches '{raw}'")

    result = {"id": a.id, "name": a.name, "asset_code": a.code,
              "code": a.code, "status": a.status}

    if mode == "lookup":
        result["message"] = "found"
    elif mode == "checkin":
        db.add(models.CheckIn(asset_id=a.id, checked_in_by=user.full_name,
                              note="Inventory check", source="barcode_scan"))
        result["message"] = "checked in"
    elif mode == "checkout":
        dept = db.get(models.Department, int(body["department_id"])) if body.get("department_id") else None
        if not dept:
            raise HTTPException(400, "Check-out mode needs a borrowing department.")
        a.checked_out = True
        a.holder_department_id = dept.id
        a.holder_person = (body.get("person") or "").strip() or None
        a.checked_out_at = models.utcnow()
        a.due_back = _parse_date(body.get("due_back"))
        db.add(models.CustodyEvent(asset_id=a.id, action="out", department_id=dept.id,
                                   person=a.holder_person, due_back=a.due_back,
                                   note="scanned out", by_user=user.email))
        result["message"] = f"checked out to {dept.name}"
    elif mode == "return":
        if not a.checked_out:
            result["message"] = "was already available"
        else:
            db.add(models.CustodyEvent(asset_id=a.id, action="in",
                                       department_id=a.holder_department_id,
                                       person=a.holder_person, note="scanned in", by_user=user.email))
            a.checked_out = False; a.holder_department_id = None
            a.holder_person = None; a.checked_out_at = None; a.due_back = None
            result["message"] = "checked in (returned)"
    else:
        raise HTTPException(400, f"Unknown scan mode '{mode}'")

    db.commit()
    result["checked_out"] = bool(a.checked_out)
    result["holder_department"] = a.holder_department.name if a.holder_department else None
    return result


@router.post("/inventory-scan")
def inventory_scan(body: dict, db: Session = Depends(get_db), user=Depends(current_user)):
    """
    Fast barcode-driven inventory checks from the dashboard. Any signed-in
    user (admin or viewer) can log one — this mirrors the already-public QR
    check-in in spirit (it only logs a scan, it doesn't edit the asset), just
    reachable from inside the dashboard instead of a phone camera.
    """
    raw = (body.get("code") or "").strip().upper()
    if not raw:
        raise HTTPException(400, "No code provided")
    a = db.query(models.Asset).filter(models.Asset.code == raw).first()
    if not a or not a.active:
        raise HTTPException(404, f"No active asset matches code '{raw}'")
    c = models.CheckIn(asset_id=a.id, checked_in_by=user.full_name,
                      note="Inventory check", source="barcode_scan")
    db.add(c); db.commit(); db.refresh(c)
    return {"id": a.id, "name": a.name, "asset_code": a.code, "code": a.code,
            "status": a.status, "ts": c.ts.isoformat(timespec="seconds")}